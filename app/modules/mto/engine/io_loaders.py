import openpyxl
import pandas as pd

# en-tete SAP source (avec caractere casse possible) -> cle canonique.
# La cle de matching ignore accents/casse/caractere de remplacement.
CANONICAL_MAP = {
    "article": "article",
    "designation": "designation",
    "texte designation article": "designation_long",
    "unite": "unite",
    "stock ul": "stock_ul",
    "stock cde": "stock_cde",
    "stock transit": "stock_transit",
    "stock total transit ul": "stock_transit_ul",
    "stock bloque": "stock_bloque",
    "stock mort": "stock_mort",
    "stock ul sans stock mort": "stock_ul_hors_mort",
    # controle qualite : capture si l'export fournit une colonne dediee
    "stock controle qualite": "stock_cq",
    "stock cq": "stock_cq",
    "controle qualite": "stock_cq",
    "stock qualite": "stock_cq",
    "stock inspection": "stock_cq",
    # variante export "BASE ARTICLE CM" (stock par emplacement)
    "utilis libre": "stock_ul_hors_mort",
    "en ctrle qual": "stock_cq",
    "transit transft": "stock_transit",
    "bloque": "stock_bloque",
    "uq": "unite",
    "mag": "mag",          # magasin (picking)
    "emplacemt": "emplacement",
    "prix unitaire": "prix_unitaire",
    "fabricant": "fabricant",
    "no de piece fabricant": "ref_fabricant",
    "groupe": "groupe",
    "hierarchie produits": "hier_produits",
    "description hierarchie pdt": "hier_pdt_desc",
    "description hierarchie niv 1": "hier_niv1_desc",
    "subsitution ca": "subst_ca",
    "subsitution designation": "subst_designation",
}

def _canon_key(h):
    import unicodedata, re
    s = "" if h is None else str(h)
    s = s.replace("�", "")  # caractere de remplacement issu d'un encodage SAP casse
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("\n", " ")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _find_header_row(path, sheet=0, scan=40):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
    header_idx = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
        keys = [_canon_key(v) for v in row]
        if "article" in keys and "designation" in keys:
            header_idx = i
            break
    wb.close()
    return header_idx

_STOCK_QTY_COLS = ["stock_ul", "stock_ul_hors_mort", "stock_cde", "stock_transit",
                   "stock_transit_ul", "stock_bloque", "stock_mort", "stock_cq"]

def _find_stock_sheet(path):
    """Detecte la feuille du stock (en-tete avec 'Stock UL' / 'Utilis. libre'),
    pour importer un classeur multi-feuilles (catalogue + stock)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            keys = {_canon_key(v) for v in row}
            if keys & {"stock ul", "utilis libre", "stock ul sans stock mort"}:
                target = name
                break
        if target:
            break
    fallback = wb.sheetnames[0]
    wb.close()
    return target or fallback


def _build_locations(df):
    """Resume des magasins/emplacements ou l'article a du stock disponible (UL>0),
    pour le picking. Retourne un DataFrame (article, emplacements) ou None."""
    if "mag" not in df.columns or "stock_ul_hors_mort" not in df.columns:
        return None
    m = df[df["stock_ul_hors_mort"].fillna(0) > 0].copy()
    if not len(m):
        return None
    mag = m["mag"].fillna("").astype(str).str.strip()
    if "emplacement" in m.columns:
        empl = m["emplacement"].fillna("").astype(str).str.strip()
        m["_loc"] = (mag + "/" + empl).str.strip("/ ")
    else:
        m["_loc"] = mag
    g = m.groupby(["article", "_loc"], as_index=False)["stock_ul_hors_mort"].sum()
    g["_s"] = g["_loc"] + ": " + g["stock_ul_hors_mort"].round().astype(int).astype(str)
    out = g.groupby("article", as_index=False)["_s"].agg(" ; ".join)
    return out.rename(columns={"_s": "emplacements"})


def load_sap(path, sheet=None):
    """Charge le stock SAP en auto (detection feuille + en-tete + mapping CANONICAL_MAP)."""
    if sheet is None:
        sheet = _find_stock_sheet(path)
    header_idx = _find_header_row(path, sheet)
    df = pd.read_excel(path, sheet_name=sheet, header=header_idx, engine="openpyxl")
    rename = {}
    for col in df.columns:
        canon = CANONICAL_MAP.get(_canon_key(col))
        if canon and canon not in rename.values():
            rename[col] = canon
    return finalize_stock(df.rename(columns=rename))


def finalize_stock(df):
    """Post-traitement du stock APRES mapping vers les cles canoniques (rename auto OU
    wizard) : filtre les lignes sans article, numerise les quantites, calcule les
    emplacements (picking) puis agrege par article."""
    df = df[df["article"].notna() & (df["article"].astype(str).str.strip() != "")].copy()
    df["article"] = df["article"].astype(str).str.strip()
    qty = [c for c in _STOCK_QTY_COLS if c in df.columns]
    for c in qty:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    locs = _build_locations(df)  # picking : ou se trouve le stock dispo (AVANT agregation)
    if df["article"].duplicated().any() and qty:
        # stock par emplacement : sommer les quantites par article (1er pour le texte)
        agg = {c: "sum" for c in qty}
        for c in df.columns:
            if c != "article" and c not in qty:
                agg[c] = "first"
        df = df.groupby("article", as_index=False).agg(agg)
    else:
        df = df.drop_duplicates(subset="article", keep="last")
    if locs is not None:
        df = df.merge(locs, on="article", how="left")
    # ne garder que les colonnes canoniques utiles (evite les colonnes brutes en doublon
    # comme 'Mag' vide qui entre en conflit avec 'mag' dans SQLite, insensible a la casse)
    keep = (set(CANONICAL_MAP.values()) - {"mag", "emplacement"}) | {"emplacements"}
    df = df[[c for c in df.columns if c in keep]]
    return df.reset_index(drop=True)

def load_mto(path):
    df = pd.read_excel(path, sheet_name="Par_item_SPEC", engine="openpyxl")
    df = df.rename(columns={
        "SPEC": "spec", "DESCRIPTION": "description", "DIAMETER": "diameter",
        "TOTAL_QTY": "total_qty", "TOTAL_PIPE_LENGTH": "total_pipe_length",
    })
    df = df[df["description"].notna()]
    return df.reset_index(drop=True)


# --- Mapping MTO multi-template : un MTO peut avoir des feuilles/colonnes variables ---

# champ cible -> indices (token, poids). Le plus specifique gagne en cas d'ambiguite.
_MTO_TARGETS = {
    "description": [("designation", 3), ("descript", 3), ("libell", 3), ("design", 3),
                    ("article", 1)],
    "diameter": [("diametre", 3), ("diam", 3), ("dn", 3), ("size", 2), ("dimension", 2)],
    "total_qty": [("qte", 3), ("quant", 3), ("besoin", 3), ("qty", 3), ("nombre", 2)],
    "spec": [("specification", 3), ("spec", 3), ("classe", 3)],
    "code_article": [("code", 3), ("sap", 3), ("material", 2), ("matricule", 2), ("reference", 2)],
    # tracabilite (regroupement / detail) : un MTO detaille (BOM) les fournit
    "line_num": [("line num", 4), ("ligne", 3), ("line", 3), ("repere ligne", 3), ("iso", 2)],
    "mark": [("mark", 3), ("item", 2), ("repere", 2), ("pos", 2)],
    "tag": [("tag", 3), ("equipement", 2), ("equipment", 2)],
    "length": [("pipe length", 4), ("longueur", 3), ("length", 3), ("long", 2)],
}

def _norm_col(c):
    import unicodedata, re
    s = unicodedata.normalize("NFKD", str(c)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def suggest_mto_mapping(columns):
    """Heuristique : associe chaque champ cible (description/diameter/total_qty/spec/
    code_article) a la colonne la plus probable. Retourne {cible: colonne|None}."""
    result = {t: None for t in _MTO_TARGETS}
    best_score = {t: 0 for t in _MTO_TARGETS}
    for c in columns:
        nc = _norm_col(c)
        scores = {t: max((w for h, w in hints if h in nc), default=0)
                  for t, hints in _MTO_TARGETS.items()}
        best_t = max(scores, key=lambda t: scores[t])
        if scores[best_t] > 0 and scores[best_t] > best_score[best_t]:
            result[best_t] = c
            best_score[best_t] = scores[best_t]
    return result


def suggest_mapping(kind, columns):
    """Pre-remplit {champ_cible: colonne} pour le wizard, par correspondance connue :
    CANONICAL_MAP (stock) / _CATALOGUE_MAP (catalogue) exactes, heuristique pour le MTO.
    Limite aux champs du schema `kind`."""
    from app.modules.mto.engine import import_schema
    targets = {k for k, _label, _req in import_schema.fields(kind)}
    if kind == "mto":
        return {t: c for t, c in suggest_mto_mapping(columns).items() if t in targets and c}
    if kind == "catalogue":
        from app.modules.mto.engine.catalogue import _CATALOGUE_MAP as cmap
    else:
        cmap = CANONICAL_MAP
    result = {}
    for col in columns:
        canon = cmap.get(_canon_key(col))
        if canon in targets and canon not in result:
            result[canon] = col
    return result

def list_mto_sheets(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets

def read_sheet_columns(path, sheet):
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", nrows=0)
    return [str(c) for c in df.columns]

def load_mto_mapped(path, sheet, mapping):
    """Charge une feuille MTO en renommant les colonnes selon `mapping`
    ({cible: colonne_source}). Produit un DataFrame aux cles canoniques."""
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    out = pd.DataFrame()
    for target, col in mapping.items():
        if col and col in df.columns:
            out[target] = df[col].values
    return finalize_mto(out, first_data_row=2)


def finalize_mto(df, first_data_row=2):
    """Post-traitement MTO apres mapping : ajoute le numero de ligne fichier (tracabilite)
    et retire les lignes sans description. `first_data_row` = 1ere ligne de donnees
    (= header_row + 2 cote wizard ; 2 quand l'en-tete est en ligne 1)."""
    out = df.copy()
    out["_row"] = range(first_data_row, first_data_row + len(out))
    if "description" in out.columns:
        out = out[out["description"].notna()]
    return out.reset_index(drop=True)


# ---------- import interactif (wizard) : lecture brute + application d'un mapping ----------

def read_raw(path, sheet=0, header_row=0, nrows=None):
    """Lit une feuille brute (colonnes telles quelles), en-tete a `header_row` (0-indexe).
    `nrows` limite les lignes (apercu du wizard). 1ere ligne de donnees = header_row + 1."""
    return pd.read_excel(path, sheet_name=sheet, header=header_row, nrows=nrows, engine="openpyxl")


def _txt(series):
    """Series -> texte propre (NaN -> '') pour la fusion de colonnes."""
    return series.astype(object).where(series.notna(), "").astype(str)


def apply_mapping(df_raw, config):
    """Applique un mapping interactif a un DataFrame brut. `config` :
        {'mapping':    {champ_cible: colonne_source},
         'transforms': {champ_cible: [{'type': nom, 'params': {...}}, ...]}}
    Transfo speciale 'concat' (params {'column': autre_colonne, 'sep': ' '}) : fusionne
    deux colonnes du fichier. Retourne un DataFrame aux cles canoniques.
    """
    from app.modules.mto.engine import transforms
    mapping = config.get("mapping", {})
    tdefs = config.get("transforms", {})
    out = pd.DataFrame(index=df_raw.index)
    for target, src in mapping.items():
        if not src or src not in df_raw.columns:
            continue
        series = df_raw[src]
        for step in tdefs.get(target, []):
            if step.get("type") == "concat":
                p = step.get("params") or {}
                col = p.get("column")
                if col in df_raw.columns:
                    series = _txt(series).str.cat(_txt(df_raw[col]), sep=p.get("sep", " "))
            else:
                series = transforms.apply_transform(series, step.get("type"), step.get("params"))
        out[target] = series.values
    return out.reset_index(drop=True)
